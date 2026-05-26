import asyncio
import logging
import requests
import re
import time
import io
from collections import defaultdict, deque
from datetime import datetime
from aiogram import Bot, Dispatcher
from aiogram.filters import CommandStart, Command
from aiogram.types import Message, BufferedInputFile
import os
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


load_dotenv()
TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
BACKEND_URL = os.getenv("BACKEND_URL")

ADMIN_IDS = [int(x) for x in os.getenv("ADMIN_IDS", "").split(",") if x]

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("bot.log", encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
logger = logging.getLogger("bot")

RATE_LIMIT_MESSAGES = 10
RATE_LIMIT_WINDOW   = 60
BAN_DURATION        = 60
MAX_MESSAGE_LENGTH  = 500

_user_history: dict[int, deque] = defaultdict(lambda: deque(maxlen=RATE_LIMIT_MESSAGES))
_banned_users: dict[int, float] = {}


def is_banned(user_id: int) -> bool:
    until = _banned_users.get(user_id)
    if until is None:
        return False
    if time.time() > until:
        del _banned_users[user_id]
        return False
    return True


def ban_user(user_id: int, username: str | None) -> None:
    _banned_users[user_id] = time.time() + BAN_DURATION
    logger.warning(
        "RATE LIMIT EXCEEDED: user_id=%s username=%s banned for %ds",
        user_id, username, BAN_DURATION
    )


def check_rate_limit(user_id: int, username: str | None) -> tuple[bool, int]:
    if is_banned(user_id):
        seconds_left = int(_banned_users[user_id] - time.time())
        return False, seconds_left

    history = _user_history[user_id]
    now = time.time()

    while history and now - history[0] > RATE_LIMIT_WINDOW:
        history.popleft()

    if len(history) >= RATE_LIMIT_MESSAGES:
        ban_user(user_id, username)
        return False, BAN_DURATION

    history.append(now)
    return True, 0


def validate_message(text: str | None) -> tuple[bool, str]:
    if not text:
        return False, "Порожнє повідомлення."
    if len(text) > MAX_MESSAGE_LENGTH:
        return False, f"Повідомлення задовге (максимум {MAX_MESSAGE_LENGTH} символів)."
    return True, ""


_summary_cache: dict = {
    "data": None,
    "text": None,
    "timestamp": 0.0,
}
CACHE_TTL = 600


def _get_report_key(data: dict) -> tuple:
    return (
        data.get("total_events"),
        data.get("joins"),
        data.get("leaves"),
        data.get("unique_players"),
    )


def _is_cache_valid(current_data: dict) -> bool:
    if _summary_cache["text"] is None:
        return False
    if time.time() - _summary_cache["timestamp"] > CACHE_TTL:
        return False
    if _summary_cache["data"] is None:
        return False
    return _get_report_key(current_data) == _get_report_key(_summary_cache["data"])


def _update_cache(data: dict, text: str) -> None:
    _summary_cache["data"] = data
    _summary_cache["text"] = text
    _summary_cache["timestamp"] = time.time()


def build_excel_report(report: dict, players: list, events: list) -> bytes:
    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2563EB")
    center      = Alignment(horizontal="center", vertical="center")
    thin        = Side(style="thin", color="D1D5DB")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill    = PatternFill("solid", fgColor="F8FAFC")

    def style_header_row(ws, row, cols):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

    def style_data_row(ws, row, cols, alternate=False):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if alternate:
                cell.fill = alt_fill

    ws1 = wb.active
    ws1.title = "Статистика"

    ws1.merge_cells("A1:B1")
    title_cell = ws1["A1"]
    title_cell.value     = f"Звіт активності сервера — {datetime.now().strftime('%d.%m.%Y')}"
    title_cell.font      = Font(bold=True, size=13)
    title_cell.alignment = center
    ws1.row_dimensions[1].height = 28

    rows = [
        ("Показник", "Значення"),
        ("Подій всього",    report.get("total_events", 0)),
        ("Входів (join)",   report.get("joins", 0)),
        ("Виходів (leave)", report.get("leaves", 0)),
        ("Унікальні гравці", report.get("unique_players", 0)),
        ("Статус",          report.get("status", "—")),
        ("Тренд",           report.get("trend", "—")),
        ("Пік активності",  f"{report.get('peak_hour')}:00" if report.get("peak_hour") is not None else "—"),
        ("Висновок",        report.get("insight", "—")),
    ]

    for i, (label, value) in enumerate(rows):
        ws1.cell(row=i+2, column=1, value=label)
        ws1.cell(row=i+2, column=2, value=value)
        if i == 0:
            style_header_row(ws1, i+2, 2)
        else:
            style_data_row(ws1, i+2, 2, alternate=(i % 2 == 0))

    ws1.column_dimensions["A"].width = 24
    ws1.column_dimensions["B"].width = 28

    ws2 = wb.create_sheet("Топ гравців")

    headers = ["#", "Гравець", "Сесій", "Перший вхід", "Останній вхід"]
    for col, h in enumerate(headers, 1):
        ws2.cell(row=1, column=col, value=h)
    style_header_row(ws2, 1, len(headers))
    ws2.row_dimensions[1].height = 22

    for i, p in enumerate(players, 1):
        ws2.cell(row=i+1, column=1, value=i)
        ws2.cell(row=i+1, column=2, value=p.get("player_name") or "—")
        ws2.cell(row=i+1, column=3, value=p.get("total_sessions", 0))
        ws2.cell(row=i+1, column=4, value=str(p.get("first_seen", "—"))[:16])
        ws2.cell(row=i+1, column=5, value=str(p.get("last_seen",  "—"))[:16])
        style_data_row(ws2, i+1, len(headers), alternate=(i % 2 == 0))

    for col, width in zip(range(1, 6), [5, 20, 10, 18, 18]):
        ws2.column_dimensions[get_column_letter(col)].width = width

    ws3 = wb.create_sheet("Останні події")

    headers3 = ["ID", "Тип події", "Гравець", "Час"]
    for col, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=col, value=h)
    style_header_row(ws3, 1, len(headers3))
    ws3.row_dimensions[1].height = 22

    event_labels = {"player_join": "Вхід", "player_leave": "Вихід"}

    for i, e in enumerate(events, 1):
        ws3.cell(row=i+1, column=1, value=e.get("id"))
        ws3.cell(row=i+1, column=2, value=event_labels.get(e.get("event_type"), e.get("event_type")))
        ws3.cell(row=i+1, column=3, value=e.get("player_name") or "—")
        ws3.cell(row=i+1, column=4, value=str(e.get("created_at", "—"))[:16])
        style_data_row(ws3, i+1, len(headers3), alternate=(i % 2 == 0))

    for col, width in zip(range(1, 5), [6, 14, 20, 18]):
        ws3.column_dimensions[get_column_letter(col)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


bot = Bot(token=TOKEN)
dp = Dispatcher()


async def guard(message: Message) -> bool:
    user_id  = message.from_user.id
    username = message.from_user.username

    if user_id not in ADMIN_IDS:
        allowed, seconds_left = check_rate_limit(user_id, username)
        if not allowed:
            await message.answer(
                f"⛔ Перевищено ліміт запитів. Спробуй через {seconds_left} с."
            )
            return False

    ok, error = validate_message(message.text)
    if not ok:
        await message.answer(f"⚠️ {error}")
        return False

    return True


@dp.message(CommandStart())
async def start(message: Message):
    if not await guard(message):
        return
    await message.answer(
        "Привіт! 🤖\n\n"
        "Я бот для аналітики ігрового сервера.\n\n"
        "Доступні команди:\n"
        "• /report — статистика за сьогодні\n"
        "• /summary — AI-аналітика (тільки адмін)\n"
        "• /table — Excel-звіт (тільки адмін)\n\n"
        "Або напиши будь-яке питання — відповім."
    )


@dp.message(Command("report"))
async def report(message: Message):
    if not await guard(message):
        return
    try:
        response = requests.get(f"{BACKEND_URL}/report/today", timeout=5)
        response.raise_for_status()
        data = response.json()

        text = (
            "📊 Звіт за сьогодні:\n\n"
            f"• Подій всього: {data.get('total_events', 0)}\n"
            f"• Join: {data.get('joins', 0)}\n"
            f"• Leave: {data.get('leaves', 0)}\n"
            f"• Унікальні гравці: {data.get('unique_players', 0)}\n"
            f"• Статус: {data.get('status', '—')}\n"
            f"• Тренд: {data.get('trend', '—')}"
        )

        await message.answer(text)

    except Exception as e:
        logger.error("REPORT ERROR: %s", e)
        await message.answer("❌ Помилка отримання звіту.")


@dp.message(Command("summary"))
async def summary(message: Message):
    if message.from_user.id not in ADMIN_IDS:
        await message.answer("⛔ У тебе немає доступу до цієї команди.")
        return

    try:
        report_resp = requests.get(f"{BACKEND_URL}/report/today", timeout=5)
        report_resp.raise_for_status()
        current_data = report_resp.json()

        if _is_cache_valid(current_data):
            await message.answer(
                "🧠 AI-аналітика (з кешу):\n\n" + _summary_cache["text"]
            )
            return

        response = requests.get(f"{BACKEND_URL}/summary/today", timeout=60)
        response.raise_for_status()
        data = response.json()

        summary_text = data.get("summary", "Немає даних")
        summary_text = re.sub(r'[#*`_>\-]', '', summary_text)
        summary_text = re.sub(r'\n\s*\n', '\n\n', summary_text)
        summary_text = summary_text[:4000]

        _update_cache(current_data, summary_text)

        await message.answer("🧠 AI-аналітика:\n\n" + summary_text)

    except Exception as e:
        logger.error("SUMMARY ERROR: %s", e)
        await message.answer("❌ Помилка AI-аналітики.")


@dp.message(Command("table"))
async def table(message: Message):
    if message.from_user.id not in ADMIN_IDS:
        await message.answer("⛔ У тебе немає доступу до цієї команди.")
        return

    await message.answer("⏳ Генерую Excel-звіт...")

    try:
        report_resp  = requests.get(f"{BACKEND_URL}/report/today",  timeout=5)
        players_resp = requests.get(f"{BACKEND_URL}/players/top",    timeout=5)
        events_resp  = requests.get(f"{BACKEND_URL}/events/recent",  timeout=5)

        report_resp.raise_for_status()
        players_resp.raise_for_status()
        events_resp.raise_for_status()

        report_data  = report_resp.json()
        players_data = players_resp.json().get("players", [])
        events_data  = events_resp.json().get("events", [])

        excel_bytes = build_excel_report(report_data, players_data, events_data)

        filename = f"report_{datetime.now().strftime('%d-%m-%Y')}.xlsx"

        await message.answer_document(
            document=BufferedInputFile(excel_bytes, filename=filename),
            caption=f"📊 Звіт активності сервера за {datetime.now().strftime('%d.%m.%Y')}"
        )

    except Exception as e:
        logger.error("TABLE ERROR: %s", e)
        await message.answer("❌ Помилка генерації звіту.")


@dp.message()
async def chat_handler(message: Message):
    if not await guard(message):
        return
    try:
        response = requests.post(
            f"{BACKEND_URL}/chat",
            json={"message": message.text},
            timeout=30
        )
        response.raise_for_status()
        data = response.json()

        reply = data.get("reply", "Немає відповіді.")
        reply = re.sub(r'[#*`_>\-]', '', reply)
        reply = reply[:4000]

        await message.answer(reply)

    except Exception as e:
        logger.error("CHAT ERROR: %s", e)
        await message.answer("❌ Помилка з'єднання з сервером.")


async def main():
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())